import openpyxl
import pandas as pd
from pathlib import Path

XLSX_PATH = Path(__file__).parent / "Cerner Benchmark Patient Summary Evaluation Sheet.xlsx"

EVALUATOR_GROUPS = {
    "Evaluator 1": "A", "Evaluator 2": "A",
    "Evaluator 3": "B", "Evaluator 4": "B",
    "Evaluator 5": "C", "Evaluator 6": "C",
    "Tester": "A",
}


def load_evaluation_data(xlsx_path: str = None) -> pd.DataFrame:
    """Parse the XLSX file including hyperlinks from Column C."""
    if xlsx_path is None:
        xlsx_path = str(XLSX_PATH)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['pvs_eval']

    rows = []
    for row_num in range(2, ws.max_row + 1):
        doc_id = ws[f'A{row_num}'].value
        if not doc_id:
            continue

        # Extract hyperlink from Column C
        cell_c = ws[f'C{row_num}']
        gdrive_url = cell_c.hyperlink.target if cell_c.hyperlink else ""

        rows.append({
            'document_id': str(doc_id),
            'org_id': str(ws[f'B{row_num}'].value or ''),
            'doc_link_text': str(cell_c.value or ''),
            'gdrive_url': gdrive_url,
            'note_type': str(ws[f'D{row_num}'].value or ''),
            'reader_specialty': str(ws[f'E{row_num}'].value or ''),
            'model_1_summary': str(ws[f'F{row_num}'].value or ''),
            'model_2_summary': str(ws[f'G{row_num}'].value or ''),
            'model_3_summary': str(ws[f'H{row_num}'].value or ''),
        })

    wb.close()
    return pd.DataFrame(rows)


def create_assignments(df: pd.DataFrame) -> dict:
    """Create evaluator group assignments. 100 notes split across 3 groups (34/33/33)."""
    total = len(df)
    split1 = total // 3          # 33
    split2 = 2 * (total // 3)    # 66

    # Group A: first 34, Group B: next 33, Group C: last 33
    group_notes = {
        "A": df.iloc[:split1 + 1].to_dict('records'),    # 34 notes (0..33)
        "B": df.iloc[split1 + 1:split2 + 1].to_dict('records'),  # 33 notes (34..66)
        "C": df.iloc[split2 + 1:].to_dict('records'),    # 33 notes (67..99)
    }

    evaluator_assignments = {}
    for evaluator, group in EVALUATOR_GROUPS.items():
        evaluator_assignments[evaluator] = group_notes[group]

    return evaluator_assignments


def get_evaluator_notes(evaluator: str, df: pd.DataFrame) -> list:
    """Get the list of note dicts assigned to a specific evaluator."""
    assignments = create_assignments(df)
    return assignments.get(evaluator, [])
